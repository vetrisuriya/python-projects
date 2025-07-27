import openpyxl

work_book = openpyxl.load_workbook("tamil-nadu-population-stat.xlsx")
# list all sheet names
sheet_names = work_book.sheetnames
sheet_lists = {}
if len(sheet_names):
    print("\nSelect Sheet:")
    for key, sheet in enumerate(sheet_names):
        sheet_lists[key + 1] = sheet
        print(f"{key + 1}. {sheet_lists[key + 1]}")
    print("0. Exit")

    choice = input(f"Enter sheet number (1-{len(sheet_lists)}): ")
    if choice != '0':
        choice = int(choice)
        sheet_name = sheet_lists.get(choice, "empty")
        if sheet_name != "empty":
            work_sheet = work_book[sheet_name]
            # maximum now of rows
            ws_total_rows = work_sheet.max_row
            ws_total_cols = work_sheet.max_column
            print("Rows:", ws_total_rows, "Columns:", ws_total_cols)

            cell_width = []
            for row in range(0, ws_total_rows):
                for col in range(0, ws_total_cols):
                    cell = work_sheet.cell(row + 1, col + 1)
                    value = str(cell.value)
                    val_len = len(value)
                    if row != 0:
                        if cell_width[col] < val_len:
                            cell_width[col] = val_len
                    else:
                        cell_width.append(val_len)

            for row in range(0, ws_total_rows):
                for col in range(0, ws_total_cols):
                    cell = work_sheet.cell(row + 1, col + 1)
                    value = str(cell.value)
                    val_len = len(value)
                    if cell_width[col] < val_len:
                        count_len = val_len - cell_width[col]
                    else:
                        count_len = cell_width[col] - val_len
                    print("| ", (" " * count_len) + value, " ", end="")
                print("")
    else:
        exit()
else:
    print("\nThere is no sheet in the workbook!")
